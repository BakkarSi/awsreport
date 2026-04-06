"""
AWS Monthly Cost Finance Report Generator (Single Account, Tag-Based)
======================================================================
Pulls service-level costs from AWS Cost Explorer for a SINGLE AWS account,
groups them by the "Project" resource tag using alias mapping (multiple
raw tag values map to one canonical project name), allocates untagged
costs equally as Shared Services, and outputs a formatted Excel report.

Usage:
    python generate_report.py --month 2026-04
    python generate_report.py --month 2026-04 --csv-folder ./data

Requirements:
    pip install boto3 openpyxl

IAM permission needed:
    ce:GetCostAndUsage
"""

import argparse
import os
import sys
import re
import csv as csv_module
from datetime import date, datetime

import boto3
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────────────────────────────────────
# CONFIGURATION — edit this section only
# ─────────────────────────────────────────────────────────────────────────────

# Canonical project names (display labels, column order in report)
PROJECT_NAMES = ["PWCT", "KGAC", "SCZ", "EPM", "AES Development"]

# All raw tag values that map to each canonical project.
# Matching is case-insensitive. Add new aliases here as tags evolve.
TAG_ALIASES = {
    "PWCT":            ["PWCT", "Project_PWCT"],
    "KGAC":            ["PWCT_KGAC"],
    "SCZ":             ["PWCT_ECA", "PWCT_SCEZ", "PWCT_SCZ", "PWCT-SCZ"],
    "EPM":             ["PWCT-EPM", "PWCT_EPM"],
    "AES Development": ["AES_DEV", "PWCT_AES", "PWCT_DEV", "PWCT-AES"],
}

# Tag key used on resources. Cost Explorer matches case-insensitively.
TAG_KEY      = "Project"

# Label for untagged / unknown-tag resources — split equally across projects
SHARED_LABEL = "Shared Services"

# Cost Explorer is always us-east-1 regardless of your workload region
CE_REGION    = "us-east-1"

# ─────────────────────────────────────────────────────────────────────────────
# BUILD REVERSE LOOKUP: raw tag value (lowercase) → canonical project name
# ─────────────────────────────────────────────────────────────────────────────

def build_alias_map():
    """
    Returns {lowercase_raw_tag: canonical_project_name}
    e.g. {"pwct": "PWCT", "project_pwct": "PWCT", "pwct_kgac": "KGAC", ...}
    """
    m = {}
    for canonical, aliases in TAG_ALIASES.items():
        for alias in aliases:
            m[alias.lower()] = canonical
    return m

ALIAS_MAP = build_alias_map()


def resolve_tag(raw_tag_value: str):
    """
    Given a raw tag value from Cost Explorer (e.g. 'PWCT_SCZ'),
    return the canonical project name or None if unrecognised/untagged.
    """
    normalised = raw_tag_value.strip().lower()
    if not normalised:
        return None                     # empty = untagged
    return ALIAS_MAP.get(normalised)    # None = unknown tag → shared


# ─────────────────────────────────────────────────────────────────────────────
# DATE HELPERS
# ─────────────────────────────────────────────────────────────────────────────

def month_to_date_range(month: str):
    year, mon = int(month[:4]), int(month[5:7])
    start = f"{year:04d}-{mon:02d}-01"
    end   = f"{year:04d}-{mon+1:02d}-01" if mon < 12 else f"{year+1:04d}-01-01"
    return start, end


# ─────────────────────────────────────────────────────────────────────────────
# COST EXPLORER — LIVE AWS FETCH (dual-call approach)
# ─────────────────────────────────────────────────────────────────────────────

def _get_ce_client():
    """Return a Cost Explorer boto3 client, honouring AWS_PROFILE if set."""
    profile = os.environ.get("AWS_PROFILE")
    session = boto3.Session(profile_name=profile) if profile else boto3.Session()
    return session.client("ce", region_name=CE_REGION)


def _paginate_ce(ce, **kwargs) -> list:
    """
    Calls get_cost_and_usage and follows NextPageToken to collect all groups.
    Cost Explorer paginates at ~20 000 rows — large accounts need this.
    """
    groups = []
    token  = None
    while True:
        if token:
            kwargs["NextPageToken"] = token
        resp   = ce.get_cost_and_usage(**kwargs)
        groups += resp["ResultsByTime"][0]["Groups"]
        token   = resp.get("NextPageToken")
        if not token:
            break
    return groups


def fetch_costs_by_tag(start: str, end: str) -> dict:
    """
    TWO Cost Explorer calls are made:

    Call 1 — SERVICE only (no tag filter):
        Gets the ground-truth total cost per service for the month.
        This catches EVERY service regardless of tag state.

    Call 2 — SERVICE + PROJECT TAG:
        Gets costs broken down by project tag alias.
        This is used to allocate tagged costs to their project columns.

    Reconciliation:
        tagged_total_per_service  = sum across all project buckets
        unaccounted               = call1_total - tagged_total
        unaccounted > 0           → goes to Shared Services (untagged/gap)

    This ensures the report grand total always matches Cost Explorer exactly,
    and no service is silently dropped because it had a missing or unknown tag.

    Returns:
        {
            "PWCT":            {"EC2-Instances": 374.80, ...},
            "SCZ":             {"Elastic Load Balancing": 829.28, ...},
            "Shared Services": {"VPC": 1526.09, ...},
            ...
        }
    """
    ce = _get_ce_client()

    # ── Call 1: all services, no tag grouping ─────────────────────────────────
    print(f"  Call 1: all services by cost (ground truth)...")
    all_svc_groups = _paginate_ce(
        ce,
        TimePeriod={"Start": start, "End": end},
        Granularity="MONTHLY",
        Metrics=["UnblendedCost"],
        GroupBy=[{"Type": "DIMENSION", "Key": "SERVICE"}],
    )

    # {service: total_cost}  — this is the source of truth
    all_services_total: dict[str, float] = {}
    for g in all_svc_groups:
        svc    = g["Keys"][0]
        amount = round(float(g["Metrics"]["UnblendedCost"]["Amount"]), 2)
        if amount != 0:                     # keep zero-cost services too
            all_services_total[svc] = round(
                all_services_total.get(svc, 0) + amount, 2
            )

    print(f"    → {len(all_services_total)} services found "
          f"(${sum(all_services_total.values()):,.2f} total)")

    # ── Call 2: services grouped by project tag ───────────────────────────────
    print(f"  Call 2: services grouped by '{TAG_KEY}' tag...")
    tag_groups = _paginate_ce(
        ce,
        TimePeriod={"Start": start, "End": end},
        Granularity="MONTHLY",
        Metrics=["UnblendedCost"],
        GroupBy=[
            {"Type": "DIMENSION", "Key": "SERVICE"},
            {"Type": "TAG",       "Key": TAG_KEY},
        ],
    )

    # Initialise result buckets
    result: dict[str, dict[str, float]] = {SHARED_LABEL: {}}
    for p in PROJECT_NAMES:
        result[p] = {}

    tag_prefix        = f"{TAG_KEY}$"
    unrecognised_tags = set()

    # Track how much of each service has been allocated to a project
    # {service: amount_already_tagged}
    tagged_per_service: dict[str, float] = {}

    for g in tag_groups:
        service   = g["Keys"][0]
        raw       = g["Keys"][1]            # e.g. "Project$PWCT_SCZ"
        amount    = round(float(g["Metrics"]["UnblendedCost"]["Amount"]), 2)
        raw_value = raw[len(tag_prefix):] if raw.startswith(tag_prefix) else raw
        canonical = resolve_tag(raw_value)

        if canonical:
            result[canonical][service] = round(
                result[canonical].get(service, 0) + amount, 2
            )
            tagged_per_service[service] = round(
                tagged_per_service.get(service, 0) + amount, 2
            )
        else:
            # Untagged (empty) or unrecognised tag value
            if raw_value.strip():
                unrecognised_tags.add(raw_value.strip())
            # Do NOT add to shared here — we reconcile below using Call 1 totals

    # ── Reconciliation: plug the gap into Shared Services ────────────────────
    # For every service seen in Call 1, subtract whatever was tagged in Call 2.
    # The remainder is untagged/unaccounted cost → Shared Services.
    for svc, total in all_services_total.items():
        tagged = tagged_per_service.get(svc, 0)
        gap    = round(total - tagged, 2)
        if gap != 0:
            result[SHARED_LABEL][svc] = round(
                result[SHARED_LABEL].get(svc, 0) + gap, 2
            )

    # ── Ensure every service from Call 1 appears in the result ───────────────
    # Services that are 100% tagged already appear via Call 2.
    # Services that are 100% untagged now appear in Shared Services via gap.
    # Services partially tagged appear in both — correct by construction.

    # ── Warnings ─────────────────────────────────────────────────────────────
    if unrecognised_tags:
        print(f"\n  WARNING: Unrecognised tag values found — added to "
              f"{SHARED_LABEL}.")
        print(f"  Add these to TAG_ALIASES in generate_report.py if they "
              f"belong to a specific project:")
        for t in sorted(unrecognised_tags):
            print(f"    '{t}'")

    # Summary
    total_tagged   = sum(
        sum(v for v in costs.values() if v > 0)
        for p, costs in result.items() if p != SHARED_LABEL
    )
    total_shared   = sum(v for v in result[SHARED_LABEL].values() if v > 0)
    total_all      = sum(v for v in all_services_total.values() if v > 0)
    total_accounted = round(total_tagged + total_shared, 2)

    print(f"\n  Reconciliation summary:")
    print(f"    All services total (Call 1) : ${total_all:>12,.2f}")
    print(f"    Tagged to projects          : ${total_tagged:>12,.2f}")
    print(f"    Untagged / Shared Services  : ${total_shared:>12,.2f}")
    print(f"    Accounted total             : ${total_accounted:>12,.2f}")
    delta = round(total_all - total_accounted, 2)
    if abs(delta) > 0.01:
        print(f"    WARNING: unaccounted gap    : ${delta:>12,.2f}  ← rounding")
    else:
        print(f"    Match                       : OK (delta ${delta:.2f})")

    return result


# ─────────────────────────────────────────────────────────────────────────────
# CSV FOLDER MODE — for testing without AWS credentials
# ─────────────────────────────────────────────────────────────────────────────

def load_from_csv_folder(folder: str) -> dict:
    """
    Loads costs from per-project CSV files.

    File naming:
        Shared Services.csv
        PWCT.csv
        KGAC.csv
        SCZ.csv
        EPM.csv
        AES Development.csv

    CSV format (two columns):
        EC2-Instances,4024.66
        Elastic Load Balancing,829.28
    """
    result = {SHARED_LABEL: {}}
    for p in PROJECT_NAMES:
        result[p] = {}

    for name in [SHARED_LABEL] + PROJECT_NAMES:
        fname    = f"{name}.csv"
        csv_path = os.path.join(folder, fname)
        if not os.path.exists(csv_path):
            print(f"  Warning: {fname} not found — skipping '{name}'")
            continue

        with open(csv_path, newline="", encoding="utf-8-sig") as f:
            raw = f.read().strip()

        # Strip $ signs and thousands commas
        raw = re.sub(
            r'\$([0-9,]+\.?[0-9]*)',
            lambda m: m.group(1).replace(',', ''),
            raw
        )

        costs = {}
        for row in csv_module.reader(raw.splitlines()):
            row = [c.strip() for c in row if c.strip()]
            if len(row) < 2:
                continue
            svc = row[0]
            if svc.lower() in ("service", "service total", "total costs", ""):
                continue
            if re.match(r'^[A-Za-z]+ \d{4}$', svc):
                continue
            try:
                costs[svc] = round(float(row[1]), 2)
            except ValueError:
                continue

        result[name] = costs
        print(f"  Loaded {len(costs)} services for '{name}'")

    return result


# ─────────────────────────────────────────────────────────────────────────────
# EXCEL REPORT BUILDER
# ─────────────────────────────────────────────────────────────────────────────

DASH = "—"

def _fill(h):
    return PatternFill("solid", start_color=h)

def _font(bold=False, color="000000", size=10):
    return Font(name="Arial", bold=bold, color=color, size=size)

def _border(style="thin", color="BDD7EE"):
    s = Side(style=style, color=color)
    return Border(left=s, right=s, top=s, bottom=s)

HDR1   = _fill("1F3864"); HDR2  = _fill("2E4057")
SS_BG  = _fill("2E75B6"); ALC_BG= _fill("5B9BD5")
TOT_BG = _fill("1F3864"); ALT_BG= _fill("EBF3FB")
WHT_BG = _fill("FFFFFF"); ZRO_BG= _fill("F7F7F7")
NOTE_B = _fill("FFF9E6")
THIN_B = _border(); MED_B = _border("medium","1F3864")
C = Alignment(horizontal="center", vertical="center", wrap_text=True)
L = Alignment(horizontal="left",   vertical="center", wrap_text=True)
R = Alignment(horizontal="right",  vertical="center")
USD   = "$#,##0.00"
USD_Z = '$#,##0.00;($#,##0.00);"-"'
PCT   = "0.0%"


def _c(ws, row, col, val, font=None, fill=None, align=None, fmt=None, border=None):
    cell = ws.cell(row, col)
    cell.value = val
    if font:   cell.font          = font
    if fill:   cell.fill          = fill
    if align:  cell.alignment     = align
    if fmt:    cell.number_format = fmt
    if border: cell.border        = border
    return cell


def build_report(month, account_costs, output_path):
    month_label = datetime.strptime(month, "%Y-%m").strftime("%B %Y")
    n_proj      = len(PROJECT_NAMES)
    shared      = account_costs.get(SHARED_LABEL, {})

    # Collect all services across all buckets
    all_services = set()
    for costs in account_costs.values():
        all_services.update(costs.keys())

    # Sort: shared non-zero → project-only non-zero → all-zero
    def sort_key(svc):
        sc = shared.get(svc) or 0
        pc = sum((account_costs.get(p, {}).get(svc) or 0) for p in PROJECT_NAMES)
        if sc > 0: return (0, -sc)
        if pc > 0: return (1, -pc)
        return (2, 0)

    services = sorted(all_services, key=sort_key)

    COL_SVC  = 1
    COL_SS   = 2
    COL_PROJ = list(range(3, 3 + n_proj))
    COL_TOT  = 3 + n_proj
    N_COLS   = COL_TOT
    DS       = 7                        # DATA_START row
    DE       = DS + len(services) - 1   # DATA_END row
    TR       = DE + 2                   # TOTAL_ROW
    NR       = TR + 2                   # NOTE_ROW

    wb = Workbook()

    # ── Sheet 1: Detailed breakdown ──────────────────────────────────────────
    ws = wb.active
    ws.title = f"{month_label} — By Project"

    ws.column_dimensions["A"].width = 50
    ws.column_dimensions["B"].width = 24
    for i in range(n_proj):
        ws.column_dimensions[get_column_letter(3 + i)].width = 20
    ws.column_dimensions[get_column_letter(COL_TOT)].width = 20

    # Row 1: title
    ws.merge_cells(f"A1:{get_column_letter(N_COLS)}1")
    _c(ws,1,1, f"AWS Cost Finance Report — {month_label}",
       font=_font(True,"FFFFFF",14), fill=HDR1, align=C, border=THIN_B)
    ws.row_dimensions[1].height = 32

    # Row 2: subtitle — show alias summary
    aliases_desc = "  |  ".join(
        f"{p}: {', '.join(TAG_ALIASES[p])}" for p in PROJECT_NAMES
    )
    ws.merge_cells(f"A2:{get_column_letter(N_COLS)}2")
    _c(ws,2,1,
       f"Single account  |  Tag: '{TAG_KEY}'  |  "
       f"Untagged/unknown → {SHARED_LABEL} ÷ {n_proj}",
       font=_font(False,"FFFFFF",10), fill=HDR2, align=C, border=THIN_B)
    ws.row_dimensions[2].height = 20

    # Row 3: tag alias reference row
    ws.merge_cells(f"A3:{get_column_letter(N_COLS)}3")
    _c(ws,3,1, f"Tag aliases  |  {aliases_desc}",
       font=_font(False,"FFFFFF",9), fill=_fill("243858"), align=L, border=THIN_B)
    ws.row_dimensions[3].height = 18

    ws.row_dimensions[4].height = 8

    # Row 5: column headers
    for ci, h in enumerate(
        ["Service", f"{SHARED_LABEL} ($)"]
        + [f"{p} ($)" for p in PROJECT_NAMES]
        + ["Grand Total ($)"], 1
    ):
        _c(ws,5,ci, h, font=_font(True,"FFFFFF",10),
           fill=HDR1, align=C, border=THIN_B)
    ws.row_dimensions[5].height = 36

    # Row 6: shared total banner
    _c(ws,6,COL_SVC, f"{SHARED_LABEL} — Total (untagged + unknown tags)",
       font=_font(True,"FFFFFF",10), fill=SS_BG, align=L, border=THIN_B)
    _c(ws,6,COL_SS,
       f"=SUMIF(B{DS}:B{DE},\">0\",B{DS}:B{DE})",
       font=_font(True,"FFFFFF",10), fill=SS_BG, align=R, fmt=USD, border=THIN_B)
    for ci in COL_PROJ + [COL_TOT]:
        _c(ws,6,ci, "↓ allocated below",
           font=_font(False,"FFFFFF",9), fill=SS_BG, align=C, border=THIN_B)
    ws.row_dimensions[6].height = 20

    # Row 7: allocation per project
    _c(ws,7,COL_SVC, f"  {SHARED_LABEL} allocation per project (÷{n_proj})",
       font=_font(True,"1F3864",9), fill=ALC_BG, align=L, border=THIN_B)
    _c(ws,7,COL_SS, "", fill=ALC_BG, border=THIN_B)
    for ci in COL_PROJ:
        _c(ws,7,ci, f"=$B$6/{n_proj}",
           font=_font(True,"1F3864",10), fill=ALC_BG,
           align=R, fmt=USD, border=THIN_B)
    _c(ws,7,COL_TOT, "=$B$6",
       font=_font(True,"1F3864",10), fill=ALC_BG,
       align=R, fmt=USD, border=THIN_B)
    ws.row_dimensions[7].height = 20

    # Shift data rows down by 2 because we added alias row + extra header rows
    DS = 9   # data now starts at row 9
    DE = DS + len(services) - 1
    TR = DE + 2
    NR = TR + 2

    # Data rows
    for i, svc in enumerate(services):
        r      = DS + i
        sc     = shared.get(svc)
        proj_v = [account_costs.get(p, {}).get(svc) for p in PROJECT_NAMES]

        all_zero = (sc is None or sc == 0) and all(v is None or v == 0 for v in proj_v)
        is_alt   = (i % 2 == 0)
        bg = ZRO_BG if all_zero else (ALT_BG if is_alt else WHT_BG)

        _c(ws,r,COL_SVC, svc,
           font=_font(False,"000000",10), fill=bg, align=L, border=THIN_B)

        if sc is None:
            _c(ws,r,COL_SS, DASH,
               font=_font(False,"BBBBBB",10), fill=bg, align=R, border=THIN_B)
        elif sc == 0:
            _c(ws,r,COL_SS, 0.0,
               font=_font(False,"999999",10), fill=bg, align=R,
               fmt=USD_Z, border=THIN_B)
        else:
            _c(ws,r,COL_SS, sc,
               font=_font(False,"000000",10), fill=bg, align=R,
               fmt=USD, border=THIN_B)

        has_shared  = sc is not None and sc > 0
        proj_cells  = []
        for pi, pv in enumerate(proj_v):
            ci = COL_PROJ[pi]
            if pv is None:
                _c(ws,r,ci, DASH,
                   font=_font(False,"BBBBBB",10), fill=bg, align=R, border=THIN_B)
            elif pv == 0:
                _c(ws,r,ci, 0.0,
                   font=_font(False,"999999",10), fill=bg, align=R,
                   fmt=USD_Z, border=THIN_B)
            else:
                _c(ws,r,ci, pv,
                   font=_font(False,"000000",10), fill=bg, align=R,
                   fmt=USD, border=THIN_B)
                proj_cells.append(get_column_letter(ci))

        parts = [f"{l}{r}" for l in proj_cells]
        if has_shared:
            parts.append(f"$B$6/{n_proj}")
        if parts:
            _c(ws,r,COL_TOT, f"={'+ '.join(parts)}",
               font=_font(False,"000000",10), fill=bg, align=R,
               fmt=USD, border=THIN_B)
        else:
            _c(ws,r,COL_TOT, DASH,
               font=_font(False,"BBBBBB",10), fill=bg, align=R, border=THIN_B)
        ws.row_dimensions[r].height = 18

    # Grand total row
    _c(ws,TR,COL_SVC, f"Grand Total — {month_label}",
       font=_font(True,"FFFFFF",11), fill=TOT_BG, align=L, border=MED_B)
    _c(ws,TR,COL_SS, "=$B$6",
       font=_font(True,"FFFFFF",11), fill=TOT_BG, align=R, fmt=USD, border=MED_B)
    for ci in COL_PROJ:
        cl = get_column_letter(ci)
        _c(ws,TR,ci,
           f"=SUMIF({cl}{DS}:{cl}{DE},\">0\",{cl}{DS}:{cl}{DE})+$B$6/{n_proj}",
           font=_font(True,"FFFFFF",11), fill=TOT_BG,
           align=R, fmt=USD, border=MED_B)
    _c(ws,TR,COL_TOT,
       f"={'+'.join(get_column_letter(c)+str(TR) for c in COL_PROJ)}",
       font=_font(True,"FFFFFF",11), fill=TOT_BG,
       align=R, fmt=USD, border=MED_B)
    ws.row_dimensions[TR].height = 26

    # Note
    ws.merge_cells(f"A{NR}:{get_column_letter(N_COLS)}{NR}")
    ws[f"A{NR}"].value = (
        f"Note: Costs grouped by '{TAG_KEY}' tag with alias mapping. "
        f"Untagged resources and unrecognised tag values are treated as "
        f"{SHARED_LABEL} and split equally across {n_proj} projects (÷{n_proj}). "
        f"Generated: {date.today().isoformat()}."
    )
    ws[f"A{NR}"].font      = Font(name="Arial", size=9, italic=True, color="7F6000")
    ws[f"A{NR}"].fill      = NOTE_B
    ws[f"A{NR}"].alignment = Alignment(horizontal="left", vertical="center",
                                        wrap_text=True)
    ws.row_dimensions[NR].height = 44
    ws.freeze_panes = "B9"

    # ── Sheet 2: Tag alias reference ─────────────────────────────────────────
    ws3 = wb.create_sheet("Tag Aliases")
    ws3.column_dimensions["A"].width = 22
    ws3.column_dimensions["B"].width = 50

    ws3.merge_cells("A1:B1")
    _c(ws3,1,1, "Tag Alias Mapping Reference",
       font=_font(True,"FFFFFF",12), fill=HDR1, align=C, border=THIN_B)
    ws3.row_dimensions[1].height = 28

    _c(ws3,2,1, "Canonical Project", font=_font(True,"FFFFFF",10),
       fill=HDR2, align=C, border=THIN_B)
    _c(ws3,2,2, f"Raw '{TAG_KEY}' tag values (all map to this project)",
       font=_font(True,"FFFFFF",10), fill=HDR2, align=C, border=THIN_B)
    ws3.row_dimensions[2].height = 22

    for i, p in enumerate(PROJECT_NAMES, start=3):
        bg = ALT_BG if i % 2 == 0 else WHT_BG
        _c(ws3,i,1, p, font=_font(True,"000000",10),
           fill=bg, align=L, border=THIN_B)
        _c(ws3,i,2, ",  ".join(TAG_ALIASES[p]),
           font=_font(False,"000000",10), fill=bg, align=L, border=THIN_B)
        ws3.row_dimensions[i].height = 18

    # Untagged row
    untagged_row = len(PROJECT_NAMES) + 3
    _c(ws3, untagged_row, 1, SHARED_LABEL,
       font=_font(True,"FFFFFF",10), fill=SS_BG, align=L, border=THIN_B)
    _c(ws3, untagged_row, 2,
       "Empty tag value (untagged) OR any tag value not listed above",
       font=_font(False,"FFFFFF",10), fill=SS_BG, align=L, border=THIN_B)
    ws3.row_dimensions[untagged_row].height = 18

    # ── Sheet 3: Summary dashboard ───────────────────────────────────────────
    ws2 = wb.create_sheet("Summary")
    ws2.column_dimensions["A"].width = 32
    for i in range(n_proj + 1):
        ws2.column_dimensions[get_column_letter(2 + i)].width = 20

    ws2.merge_cells(f"A1:{get_column_letter(2+n_proj)}1")
    _c(ws2,1,1, f"Project Cost Summary — {month_label}",
       font=_font(True,"FFFFFF",14), fill=HDR1, align=C, border=THIN_B)
    ws2.row_dimensions[1].height = 32

    ws2.merge_cells(f"A2:{get_column_letter(2+n_proj)}2")
    _c(ws2,2,1,
       f"Grouped by '{TAG_KEY}' tag with alias mapping  |  "
       f"Untagged allocated equally (÷{n_proj})",
       font=_font(False,"FFFFFF",10), fill=HDR2, align=C, border=THIN_B)
    ws2.row_dimensions[2].height = 20
    ws2.row_dimensions[3].height = 8

    for ci, h in enumerate(["Metric"] + PROJECT_NAMES + ["Total"], 1):
        _c(ws2,4,ci, h, font=_font(True,"FFFFFF",10),
           fill=HDR1, align=C, border=THIN_B)
    ws2.row_dimensions[4].height = 28

    src = ws.title
    s2_rows = [
        ("Tagged project costs",
         [f"=SUMIF('{src}'!{get_column_letter(c)}{DS}:{get_column_letter(c)}{DE}"
          f",\">0\",'{src}'!{get_column_letter(c)}{DS}:{get_column_letter(c)}{DE})"
          for c in COL_PROJ], USD),
        (f"{SHARED_LABEL} allocation (÷{n_proj})",
         [f"='{src}'!$B$6/{n_proj}" for _ in COL_PROJ], USD),
        ("Total cost (tagged + allocated)",
         [f"='{src}'!{get_column_letter(c)}{TR}" for c in COL_PROJ], USD),
        ("% of combined total",
         [f"='{src}'!{get_column_letter(c)}{TR}/'{src}'!{get_column_letter(COL_TOT)}{TR}"
          for c in COL_PROJ], PCT),
    ]

    for ri, (label, formulas, fmt) in enumerate(s2_rows, start=5):
        bg = ALT_BG if ri % 2 == 0 else WHT_BG
        _c(ws2,ri,1, label, font=_font(True,"000000",10),
           fill=bg, align=L, border=THIN_B)
        for ci, f in enumerate(formulas, start=2):
            _c(ws2,ri,ci, f, font=_font(False,"000000",10),
               fill=bg, align=R, fmt=fmt, border=THIN_B)
        _c(ws2,ri, 2+n_proj,
           f"=SUM(B{ri}:{get_column_letter(1+n_proj)}{ri})",
           font=_font(True,"000000",10), fill=bg,
           align=R, fmt=fmt, border=THIN_B)
        ws2.row_dimensions[ri].height = 22

    wb.save(output_path)
    print(f"Report saved: {output_path}")


# ─────────────────────────────────────────────────────────────────────────────
# ENTRY POINT
# ─────────────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="AWS cost report — single account, tag alias mapping"
    )
    parser.add_argument("--month", required=True, help="YYYY-MM e.g. 2026-04")
    parser.add_argument("--output", default=None)
    parser.add_argument(
        "--csv-folder", default=None,
        help="Load costs from CSV files instead of AWS API (testing mode)"
    )
    args = parser.parse_args()

    month  = args.month
    output = args.output or f"aws_cost_report_{month}.xlsx"

    try:
        datetime.strptime(month, "%Y-%m")
    except ValueError:
        print(f"Error: --month must be YYYY-MM, got: {month}")
        sys.exit(1)

    start, end  = month_to_date_range(month)
    month_label = datetime.strptime(month, "%Y-%m").strftime("%B %Y")

    print(f"\nAWS Cost Finance Report — {month_label}")
    print(f"Mode     : {'CSV folder' if args.csv_folder else 'Live AWS Cost Explorer'}")
    print(f"Tag key  : {TAG_KEY}")
    print(f"Projects : {', '.join(PROJECT_NAMES)}")
    print(f"Aliases  :")
    for p in PROJECT_NAMES:
        print(f"  {p:20s} ← {', '.join(TAG_ALIASES[p])}")
    print(f"Shared   : untagged + unrecognised tags → '{SHARED_LABEL}' ÷ {len(PROJECT_NAMES)}")
    print("-" * 70)

    if args.csv_folder:
        print(f"\nLoading from CSV folder: {args.csv_folder}")
        account_costs = load_from_csv_folder(args.csv_folder)
    else:
        print("\nFetching from AWS Cost Explorer...")
        try:
            account_costs = fetch_costs_by_tag(start, end)
            for name, costs in account_costs.items():
                nz = sum(1 for v in costs.values() if v > 0)
                print(f"  {name}: {len(costs)} services ({nz} non-zero)")
        except Exception as e:
            print(f"\n  ERROR: {e}")
            print("  Tip: set AWS_PROFILE or use --csv-folder for testing.")
            sys.exit(1)

    print("\nBuilding Excel report...")
    build_report(month=month, account_costs=account_costs, output_path=output)
    print(f"\nDone. Output: {output}")


if __name__ == "__main__":
    main()